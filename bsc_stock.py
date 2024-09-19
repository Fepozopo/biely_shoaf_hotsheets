import openpyxl


def bsc_stock():
    print('Updating everyday stock...')
    print("ROW | SKU | ON HAND | ON PO | ON SO/BO")

    wb1 = openpyxl.load_workbook('./xlsx/BSC-IM_StockStatus.xlsx')
    ws1 = wb1['Sheet1']

    wb3 = openpyxl.load_workbook('./xlsx/BSC_HOTSHEET.xlsx')
    ws3 = wb3['Everyday']

    sku_col = 0  # 'A' column index in ws1
    on_hand_col = 10  # 'K' column index in ws1
    on_po_col = 12  # 'M' column index in ws1
    on_so_col = 15  # 'P' column index in ws1
    on_bo_col = 19  # 'T' column index in ws1
    start_row_ws3 = 2  # First row in ws3 to check, e.g., row 2 (E2 corresponds to E2, F2, etc.)
    ws1_pointer = 1  # Start pointer for ws1

    for row_ws3 in range(start_row_ws3, ws3.max_row + 1):
        sku_ws3 = ws3[f'D{row_ws3}'].value  # SKU in column 'D' in ws3

        if sku_ws3 is None:
            continue  # Skip rows with no SKU in ws3

        for row_ws1 in range(ws1_pointer, ws1.max_row + 1): # ------- Iterate through rows in ws1 starting from ws1_pointer -------
            sku_ws1 = ws1[row_ws1][sku_col].value  # SKU in column 'A' in ws1

            if sku_ws1 is None:
                continue
            
            if sku_ws3.strip() in sku_ws1.strip():
                if ws1[row_ws1 + 2][on_hand_col].value is None:
                    # Update 'E' (on_hand), 'F' (on_po), and 'H' (on_so, on_bo) in ws3
                    ws3[f'E{row_ws3}'].value = ws1[row_ws1 + 1][on_hand_col].value
                    ws3[f'F{row_ws3}'].value = ws1[row_ws1 + 1][on_po_col].value
                    ws3[f'H{row_ws3}'].value = ws1[row_ws1 + 1][on_so_col].value + ws1[row_ws1 + 1][on_bo_col].value
                    print(row_ws3, "|", sku_ws3, "|", ws1[row_ws1 + 1][on_hand_col].value, "|", ws1[row_ws1 + 1][on_po_col].value, "|", ws1[row_ws1 + 1][on_so_col].value, "|", ws1[row_ws1 + 1][on_bo_col].value)
                else:
                    # Update 'E' (on_hand), 'F' (on_po), and 'H' (on_so, on_bo) in ws3
                    ws3[f'E{row_ws3}'].value = ws1[row_ws1 + 2][on_hand_col].value
                    ws3[f'F{row_ws3}'].value = ws1[row_ws1 + 2][on_po_col].value
                    ws3[f'H{row_ws3}'].value = ws1[row_ws1 + 2][on_so_col].value + ws1[row_ws1 + 2][on_bo_col].value
                    print(row_ws3, "|", sku_ws3, "|", ws1[row_ws1 + 2][on_hand_col].value, "|", ws1[row_ws1 + 2][on_po_col].value, "|", ws1[row_ws1 + 2][on_so_col].value, "|", ws1[row_ws1 + 2][on_bo_col].value)

                ws1_pointer = row_ws1 + 1
                break  # Move to the next row in ws3 once a match is found

    wb1.close()
    print('Saving file...')
    wb3.save('./xlsx/BSC_HOTSHEET.xlsx')
    wb3.close()


def bsc_stock_winter():
    print('Updating holiday stock...')
    print("ROW | SKU | ON HAND | ON PO | ON SO/BO")

    wb1 = openpyxl.load_workbook('./xlsx/BSC-IM_StockStatus.xlsx')
    ws1 = wb1['Sheet1']

    wb3 = openpyxl.load_workbook('./xlsx/BSC_HOTSHEET.xlsx')
    ws3 = wb3['Winter Holiday']

    sku_col = 0  # 'A' column index in ws1
    on_hand_col = 10  # 'K' column index in ws1
    on_po_col = 12  # 'M' column index in ws1
    on_so_col = 15  # 'P' column index in ws1
    on_bo_col = 19  # 'T' column index in ws1
    start_row_ws3 = 2  # First row in ws3 to check, e.g., row 2 (E2 corresponds to E2, F2, etc.)
    ws1_pointer = 1  # Start pointer for ws1

    for row_ws3 in range(start_row_ws3, ws3.max_row + 1):
        sku_ws3 = ws3[f'E{row_ws3}'].value  # SKU in column 'E' in ws3

        if sku_ws3 is None:
            continue  # Skip rows with no SKU in ws3

        for row_ws1 in range(ws1_pointer, ws1.max_row + 1): # ------- Iterate through rows in ws1 starting from ws1_pointer -------
            sku_ws1 = ws1[row_ws1][sku_col].value  # SKU in column 'A' in ws1

            if sku_ws1 is None:
                continue
            
            if sku_ws3.strip() in sku_ws1.strip():
                if ws1[row_ws1 + 2][on_hand_col].value is None:
                    # Update 'F' (on_hand), 'I' (on_po), and 'G' (on_so, on_bo) in ws3
                    ws3[f'F{row_ws3}'].value = ws1[row_ws1 + 1][on_hand_col].value
                    ws3[f'I{row_ws3}'].value = ws1[row_ws1 + 1][on_po_col].value
                    ws3[f'G{row_ws3}'].value = ws1[row_ws1 + 1][on_so_col].value + ws1[row_ws1 + 1][on_bo_col].value
                    print(row_ws3, "|", sku_ws3, "|", ws1[row_ws1 + 1][on_hand_col].value, "|", ws1[row_ws1 + 1][on_po_col].value, "|", ws1[row_ws1 + 1][on_so_col].value, "|", ws1[row_ws1 + 1][on_bo_col].value)
                else:
                    # Update 'F' (on_hand), 'I' (on_po), and 'G' (on_so, on_bo) in ws3
                    ws3[f'F{row_ws3}'].value = ws1[row_ws1 + 2][on_hand_col].value
                    ws3[f'I{row_ws3}'].value = ws1[row_ws1 + 2][on_po_col].value
                    ws3[f'G{row_ws3}'].value = ws1[row_ws1 + 2][on_so_col].value + ws1[row_ws1 + 2][on_bo_col].value
                    print(row_ws3, "|", sku_ws3, "|", ws1[row_ws1 + 2][on_hand_col].value, "|", ws1[row_ws1 + 2][on_po_col].value, "|", ws1[row_ws1 + 2][on_so_col].value, "|", ws1[row_ws1 + 2][on_bo_col].value)

                ws1_pointer = row_ws1 + 1
                break  # Move to the next row in ws3 once a match is found

    wb1.close()
    print('Saving file...')
    wb3.save('./xlsx/BSC_HOTSHEET.xlsx')
    wb3.close()