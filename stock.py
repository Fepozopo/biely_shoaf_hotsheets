import openpyxl


class Update_Stock:
    def __init__(self, hotsheet, section, report, start, sku, on_hand, on_po, on_so_bo):
        self.start = start
        self.hotsheet = hotsheet
        self.section = section
        self.report = report
        self.sku = sku
        self.on_hand = on_hand
        self.on_po = on_po
        self.on_so_bo = on_so_bo

    def update(self):
        print("ROW | SKU | ON HAND | ON PO | ON SO/BO")

        wb1 = openpyxl.load_workbook(f'{self.report}')
        ws1 = wb1['Sheet1']

        wb3 = openpyxl.load_workbook(f'{self.hotsheet}')
        ws3 = wb3[f'{self.section}']

        sku_col = 0  # 'A' column index in ws1
        on_hand_col = 10  # 'K' column index in ws1
        on_po_col = 12  # 'M' column index in ws1
        on_so_col = 15  # 'P' column index in ws1
        on_bo_col = 19  # 'T' column index in ws1
        ws1_pointer = 1  # Start pointer for ws1

        for row_ws3 in range(self.start, ws3.max_row + 1):
            sku_ws3 = ws3[f'{self.sku}{row_ws3}'].value  # SKU column in ws3

            if sku_ws3 is None:
                continue  # Skip rows with no SKU in ws3

            for row_ws1 in range(ws1_pointer, ws1.max_row + 1): # ------- Iterate through rows in ws1 starting from ws1_pointer -------
                sku_ws1 = ws1[row_ws1][sku_col].value  # SKU in column 'A' in ws1

                if sku_ws1 is None:
                    continue
                
                if sku_ws3.strip() in sku_ws1.strip():
                    if ws1[row_ws1 + 2][on_hand_col].value is None:
                        # Update (on_hand), (on_po), and (on_so, on_bo) in ws3
                        ws3[f'{self.on_hand}{row_ws3}'].value = ws1[row_ws1 + 1][on_hand_col].value
                        ws3[f'{self.on_po}{row_ws3}'].value = ws1[row_ws1 + 1][on_po_col].value
                        ws3[f'{self.on_so_bo}{row_ws3}'].value = ws1[row_ws1 + 1][on_so_col].value + ws1[row_ws1 + 1][on_bo_col].value
                        print(row_ws3, "|", sku_ws3, "|", ws1[row_ws1 + 1][on_hand_col].value, "|", ws1[row_ws1 + 1][on_po_col].value, "|", ws1[row_ws1 + 1][on_so_col].value, "|", ws1[row_ws1 + 1][on_bo_col].value)
                    else:
                        # Update (on_hand), (on_po), and (on_so, on_bo) in ws3
                        ws3[f'{self.on_hand}{row_ws3}'].value = ws1[row_ws1 + 2][on_hand_col].value
                        ws3[f'{self.on_po}{row_ws3}'].value = ws1[row_ws1 + 2][on_po_col].value
                        ws3[f'{self.on_so_bo}{row_ws3}'].value = ws1[row_ws1 + 2][on_so_col].value + ws1[row_ws1 + 2][on_bo_col].value
                        print(row_ws3, "|", sku_ws3, "|", ws1[row_ws1 + 2][on_hand_col].value, "|", ws1[row_ws1 + 2][on_po_col].value, "|", ws1[row_ws1 + 2][on_so_col].value, "|", ws1[row_ws1 + 2][on_bo_col].value)

                    ws1_pointer = row_ws1 + 1
                    break  # Move to the next row in ws3 once a match is found

        wb1.close()
        print('Saving file...')
        wb3.save(f'{self.hotsheet}')
        wb3.close()