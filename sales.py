import openpyxl


class Update_Sales():
    def __init__(self, hotsheet, section, report, start, sku, ytd):
        self.start = start
        self.hotsheet = hotsheet
        self.section = section
        self.report = report
        self.sku = sku
        self.ytd = ytd

    def update(self):
        print("ROW | SKU | YTD")

        wb2 = openpyxl.load_workbook(f'{self.report}')
        ws2 = wb2['Sheet1']

        wb3 = openpyxl.load_workbook(f'{self.hotsheet}')
        ws3 = wb3[f'{self.section}']

        sku_col = 0  # 'A' column index in ws2
        ytd_col = 18  # 'S' column index in ws2
        kit_col = 9  # 'J' column index in ws2
        ws2_pointer = 1  # Start pointer for ws2

        for row_ws3 in range(self.start, ws3.max_row + 1):
            sku_ws3 = ws3[f'{self.sku}{row_ws3}'].value  # SKU column in ws3

            if sku_ws3 is None:
                continue  # Skip rows with no SKU in ws3

            for row_ws2 in range(ws2_pointer, ws2.max_row + 1): # ------- Iterate through rows in ws2 starting from ws2_pointer -------
                sku_ws2 = ws2[row_ws2][sku_col].value  # SKU in column 'A' in ws2

                if sku_ws2 is None:
                    continue

                if sku_ws3.strip() == sku_ws2.strip():
                    if ws2[row_ws2 + 1][kit_col].value == "Kit":
                        if "20-" in sku_ws2.strip() or "21-" in sku_ws2.strip() or "22-" in sku_ws2.strip() or "24-" in sku_ws2.strip():
                            # Update (ytd) in ws3
                            ws3[f'{self.ytd}{row_ws3}'].value = ws2[row_ws2 + 2][ytd_col].value
                            print(row_ws3, "|", sku_ws3, "|", ws2[row_ws2 + 2][ytd_col].value)
                        else:
                            # Update (ytd) in ws3 * 10
                            ws3[f'{self.ytd}{row_ws3}'].value = ws2[row_ws2 + 2][ytd_col].value * 10
                            print(row_ws3, "|", sku_ws3, "|", ws2[row_ws2 + 2][ytd_col].value * 10)
                    else:
                        # Update (ytd) in ws3
                        ws3[f'{self.ytd}{row_ws3}'].value = ws2[row_ws2 + 2][ytd_col].value
                        print(row_ws3, "|", sku_ws3, "|", ws2[row_ws2 + 2][ytd_col].value)

                    ws2_pointer = row_ws2 + 1
                    break  # Move to the next row in ws3 once a match is found
                    
                

        wb2.close()
        print('Saving file...')
        wb3.save(f'{self.hotsheet}')
        wb3.close()