import openpyxl


def smd_sales():
    print('Updating everyday sales...')
    print("ROW | SKU | YTD")

    wb2 = openpyxl.load_workbook('./xlsx/SMD-IM_SalesAnalysisCondensed.xlsx')
    ws2 = wb2['Sheet1']

    wb3 = openpyxl.load_workbook('./xlsx/SMD_HOTSHEET.xlsx')
    ws3 = wb3['EVERYDAY']

    sku_col = 0  # 'A' column index in ws2
    ytd_col = 18  # 'S' column index in ws2
    start_row_ws3 = 3  # First row in ws3 to check, e.g., row 3 (E3 corresponds to E3, F3, etc.)
    ws2_pointer = 1  # Start pointer for ws2

    for row_ws3 in range(start_row_ws3, ws3.max_row + 1):
        sku_ws3 = ws3[f'E{row_ws3}'].value  # SKU in column 'E' in ws3

        if sku_ws3 is None:
            continue  # Skip rows with no SKU in ws3

        for row_ws2 in range(ws2_pointer, ws2.max_row + 1): # ------- Iterate through rows in ws2 starting from ws2_pointer -------
            sku_ws2 = ws2[row_ws2][sku_col].value  # SKU in column 'A' in ws2

            if sku_ws2 is None:
                continue

            if sku_ws3.strip() in sku_ws2.strip():
                # Update 'P' (ytd) in ws3
                ws3[f'P{row_ws3}'].value = ws2[row_ws2 + 2][ytd_col].value
                print(row_ws3, "|", sku_ws3, "|", ws2[row_ws2 + 2][ytd_col].value)

                ws2_pointer = row_ws2 + 1
                break  # Move to the next row in ws3 once a match is found
                
            

    wb2.close()
    print('Saving file...')
    wb3.save('./xlsx/SMD_HOTSHEET.xlsx')
    wb3.close()


def smd_sales_holiday():
    print('Updating holiday sales...')
    print("ROW | SKU | YTD")

    wb2 = openpyxl.load_workbook('./xlsx/SMD-IM_SalesAnalysisCondensed.xlsx')
    ws2 = wb2['Sheet1']

    wb3 = openpyxl.load_workbook('./xlsx/SMD_HOTSHEET.xlsx')
    ws3 = wb3['HOLIDAY']

    sku_col = 0  # 'A' column index in ws2
    ytd_col = 18  # 'S' column index in ws2
    start_row_ws3 = 2  # First row in ws3 to check, e.g., row 2 (E2 corresponds to E2, F2, etc.)
    ws2_pointer = 1  # Start pointer for ws1

    for row_ws3 in range(start_row_ws3, ws3.max_row + 1):
        sku_ws3 = ws3[f'C{row_ws3}'].value  # SKU in column 'C' in ws3

        if sku_ws3 is None:
            continue  # Skip rows with no SKU in ws3

        for row_ws2 in range(ws2_pointer, ws2.max_row + 1): # ------- Iterate through rows in ws2 starting from ws2_pointer -------
            sku_ws2 = ws2[row_ws2][sku_col].value  # SKU in column 'A' in ws2

            if sku_ws2 is None:
                continue

            if sku_ws3.strip() in sku_ws2.strip():
                # Update 'N' (ytd) in ws3
                ws3[f'N{row_ws3}'].value = ws2[row_ws2 + 2][ytd_col].value
                print(row_ws3, "|", sku_ws3, "|", ws2[row_ws2 + 2][ytd_col].value)

                ws2_pointer = row_ws2 + 1
                break  # Move to the next row in ws3 once a match is found

    wb2.close()
    print('Saving file...')
    wb3.save('./xlsx/SMD_HOTSHEET.xlsx')
    wb3.close()